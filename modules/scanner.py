import os
import openpyxl
from .algorithms import BinarySearch

class Scanner:
    ''' Object use to compare installed packages with an approved baseline '''
    def __init__(self, configfile):
        self.configfile = configfile
        self.baseline_directory = os.path.join(os.getcwd(), "baselines")
        self.loaded = False
        self.rolodex = {}
        self.baseline_file = ""
        self.hostname = ""
        self.load()

        # Used for metric purposes
        self.additional = []
        self.missing = []
        self.version_mismatch = []
        self.arch_mismatch = []
    
    #   
    def load(self):
        ''' Loads a baseline file for the selected config file if present '''
        # Build expected baseline filename
        self.baseline_file = os.path.join(self.baseline_directory, self.configfile.replace(".txt",".baseline.xlsx").strip())
        # Does baseline exist in baseline directory?
        if os.path.basename(self.baseline_file) in os.listdir(self.baseline_directory):
            self.loaded = True


    def build(self):
        '''
        Builds a dictionary from baseline packages used to compare with.
        Packages are distributed into alphabetic keys to speed up search
        process and are removed if the current package satisfies baseline
        requirements.
        '''
        workbook = openpyxl.load_workbook(filename=self.baseline_file)
        sheetnames = workbook.sheetnames
        for sheet in sheetnames:
            # Add each sheet as rolodex key
            if not sheet in self.rolodex.keys():
                self.rolodex[sheet] = {}
                current_sheet = workbook[sheet]
                # Get each row from baseline
                for row in current_sheet.iter_rows(values_only=True):
                    try:
                        # Package information
                        package_name = row[0].strip()
                        package_version = row[1].strip()
                        package_arch = ""

                        # Architecture only available in Linux 
                        if not "windows" in self.hostname:
                            package_arch = row[2].strip()

                        # First letter of package name for rolodex organization
                        rolodex_index = package_name[:1].upper()

                        if not rolodex_index in self.rolodex[sheet]:
                            # Add new alphabetic key 
                            self.rolodex[sheet][rolodex_index] = {"packages":{}}
                        # Add each package to its correct letter index
                        self.rolodex[sheet][rolodex_index]["packages"][package_name] = {"version":package_version, "arch":package_arch}
                    except:
                        pass


    def isLoaded(self):
        return self.loaded


    def setHostnameAndBuild(self, hostname):
        ''' Sets the hostname and builds the rolodex of baseline packages '''
        self.hostname = hostname
        self.build()

    
    def getRolodexSection(self, key, d=None):
        ''' Returns dictionary of rolodex values stored at a given key '''
        section = False
        try:
            if not d:
                section = self.rolodex[self.hostname][key]
            else:
                section = d[key]
        except Exception as e:
            print(str(e))
        return section

    
    def start(self, packages):
        ''' Compares a list packages installed on the host with the baseline '''
        for package in packages:
            try:
                package_name = ""
                package_version = ""
                package_arch = ""

                if "windows" in self.hostname:
                    package_name, package_version = package.strip().split(" ")
                else:
                    package_name, package_version, package_arch = package.strip().split(" ")
                
                idx = package_name.strip()[:1].upper()
                
                # Only proceed if the alphabetic key is present in the rolodex for this hostname
                section = self.getRolodexSection(key=idx)
                if section:

                    keys = list(section["packages"].keys())
                    # Order list for BinarySearch
                    keys.sort(key=str.lower)
                    # Package check
                    if BinarySearch(array=keys, x=package_name, start=0, end=len(keys)-1):
                        # Version check
                        if package_version == self.rolodex[self.hostname][idx]["packages"][package_name]["version"]:
                            # Architecture check
                            if package_arch == self.rolodex[self.hostname][idx]["packages"][package_name]["arch"]:
                                # Valid package found, remove from rolodex
                                del self.rolodex[self.hostname][idx]["packages"][package_name]
                            else:
                                # Architecture invalid
                                required_arch = self.rolodex[self.hostname][idx]["packages"][package_name]["arch"]
                                self.arch_mismatch.append("{} {} installed; requires: {}".format(package_name, package_version, required_arch))
                        else:
                            # Version does not match
                            required_version = self.rolodex[self.hostname][idx]["packages"][package_name]["version"]
                            self.version_mismatch.append("{} {} installed; requires: {}".format(package_name, package_version, required_version))
                    else:
                        # New package that is not in the baseline
                        self.additional.append("{}.{}.{}".format(package_name, package_version, package_arch))
                else:
                    # New package that is not in the baseline
                    # Rolodex did not include this index
                    self.additional.append("{}.{}.{}".format(package_name, package_version, package_arch))                   
            except:
                pass

        # Display findings to the user
        self.results()

        # Reset
        self.__init__(self.configfile)


    def calculate_remaining_packages(self):
        ''' Count the number of packages remaining in the rolodex '''
        i = 0
        for ascii in range(65, 90):
            try:
                idx = str(chr(ascii))
                for key, value in self.rolodex[self.hostname][idx]["packages"].items():
                    if key:
                        i += 1
                        self.missing.append("[{}, {}. {}]".format(key, value["version"], value["arch"]))
            except:
                return i
        return i


    def results(self):
        ''' Print Scanner results to the screen '''
        num_additional = len(self.additional)
        num_version = len(self.version_mismatch)
        num_arch = len(self.arch_mismatch)
        num_missing = int(self.calculate_remaining_packages())

        if num_missing > 0:
            print("\t\t\033[1;31;40mMissing Packages:")
            for package in self.missing:
                print("\t\t\t{}".format(package))

        if num_additional > 0:
            print("\t\t\033[1;31;40mAdditional Packages:")
            for package in self.additional:
                print("\t\t\t{}".format(package))

        if num_version > 0:
            print("\t\t\033[1;31;40mVersion Mismatch:")
            for package in self.version_mismatch:
                print("\t\t\t{}".format(package))

        if num_arch > 0:
            print("\t\t\033[1;31;40mArch Mismatch:")
            for package in self.arch_mismatch:
                print("\t\t\t{}".format(package))